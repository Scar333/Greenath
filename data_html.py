import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook


def data_html(html_data):
    html, html_2 = html_data[0],html_data[-1]

    # Парсим таблицу и записываем ее
    tables = pd.read_html(html)
    df = pd.DataFrame(tables[3])

    pd.ExcelWriter('Table.xlsx')

    df.drop(df.columns[[1, 2]], axis=1, inplace=True, errors='ignore')
    df.drop(index=[1])
    df.to_excel('./Table.xlsx')

    # Чистим от индексов т.к. мультииндекс в Excel не убирается
    rowsList = [1]
    wb = Workbook('./Table.xlsx')
    ws = wb.active
    wb = load_workbook('./Table.xlsx')
    currentSheet = wb.worksheets[0]
    for i in reversed(rowsList):
        currentSheet.delete_rows(i)
        currentSheet.delete_cols(i)
    for i in reversed(rowsList):
        currentSheet.delete_rows(i)
    wb.save('./Table.xlsx')

    # Переименовывваем столбцы
    file_home = 'Table.xlsx'
    wb = load_workbook(filename= file_home)
    sheet_ranges = wb['Sheet1']
    ws = wb ['Sheet1']                          # Получить лист в соответствии с именем листа Sheet1
    ws ["A1"] = 'Дата USD/RUB'
    ws ["B1"] = 'Курс USD/RUB'
    ws ["C1"] = 'Время USD/RUB'
    sheet_ranges.unmerge_cells('C1:D1')
    wb.save('Table.xlsx')

    # Парсим вторую таблицу и записываем ее
    tables = pd.read_html(html_2)    
    df = pd.DataFrame(tables[3])

    pd.ExcelWriter('Table_2.xlsx')

    df.drop(df.columns[[1, 2]], axis=1, inplace=True, errors='ignore')
    df.drop(index=[1])
    df.to_excel('./Table_2.xlsx')

    # Чистим от индексов т.к. мультииндекс в Excel не убирается
    rowsList = [1]
    wb = Workbook('./Table_2.xlsx')
    ws = wb.active
    wb = load_workbook('./Table_2.xlsx')
    currentSheet = wb.worksheets[0]
    for i in reversed(rowsList):
        currentSheet.delete_rows(i)
        currentSheet.delete_cols(i)
    for i in reversed(rowsList):
        currentSheet.delete_rows(i)
    wb.save('./Table_2.xlsx')

    # Переименовывваем столбцы
    file_home = 'Table_2.xlsx'
    wb = load_workbook(filename= file_home)
    sheet_ranges = wb['Sheet1']
    ws = wb ['Sheet1']                          # Получить лист в соответствии с именем листа Sheet1
    ws ["A1"] = 'Дата JPY/RUB'
    ws ["B1"] = 'Курс JPY/RUB'
    ws ["C1"] = 'Время JPY/RUB'
    sheet_ranges.unmerge_cells('C1:D1')
    wb.save('Table_2.xlsx')


# Объеденение таблиц
def data_merger():
    # Читаем и объединяем их
    df1 = pd.read_excel('Table.xlsx')
    df2 = pd.read_excel('Table_2.xlsx')
    new_tab = df1.join(df2)
    new_tab.to_excel('New.xlsx')

    # Читаем и проводим математические действия
    df3 = pd.read_excel('New.xlsx')
    df3['Результат'] = df3['Курс USD/RUB'] / df3['Курс JPY/RUB']

    # Замена/Переименовывание столбцов местами
    df3['E'] = df3['Курс JPY/RUB'] / 100000
    df_new = df3.reindex(columns=['Дата USD/RUB', 'Курс USD/RUB', 'Время USD/RUB', 'Дата JPY/RUB', 'E', 'Время JPY/RUB', 'Результат', 'Курс JPY/RUB'])
    df_new.rename(columns={'Курс JPY/RUB':'sdaasd'}, inplace=True)
    df_new.rename(columns={'E':'Курс JPY/RUB'}, inplace=True)
    df_new.drop(columns=['sdaasd', 'Результат'], axis=1, inplace=True)
    df_new['Результат'] = df3['Курс USD/RUB'] / df3['Курс JPY/RUB']
    df_new.to_excel('New.xlsx')

    # Количество строк в файле
    global size
    size = len(df_new.index) 

    # Чистка столбцов
    rowsList = [1]
    wb = Workbook('./New.xlsx')
    ws = wb.active
    wb = load_workbook('./New.xlsx')
    currentSheet = wb.worksheets[0]
    for i in reversed(rowsList):
        currentSheet.delete_cols(i)
    
    # Получить лист в соответствии с именем листа Sheet1
    ws = wb ['Sheet1']

    # Автоширина столбцов
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))  
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.save('./New.xlsx')

    return size